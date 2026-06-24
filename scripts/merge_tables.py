#!/usr/bin/env python3
"""Fully merge the two plural tables into one master sheet.

Each (singular pattern, plural pattern, type) correspondence becomes a single
row. Near-duplicate rows that the two source tables wrote slightly differently
(a missing sukun/fatha, or an اسم/صفة label mismatch) are reconciled into one
row, and a الفروق column explains whatever genuinely differs between the two
source tables for that row.

Output columns:
  المفرد | النوع | الجمع | الأمثلة | الفروق
"""
import csv
import os
import re
from collections import Counter

import openpyxl
from openpyxl.styles import Alignment, Font

# archival: reads the two extracted source tables from sources/ and writes the
# pre-revision unified table there too. The active pipeline is build_dataset.py.
ROOT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "sources")

RLM = "‏"  # right-to-left mark
SUKUN, FATHA, TATWEEL = "ْ", "َ", "ـ"

S2P_NAME = "من المفرد إلى الجمع"
P2S_NAME = "من الجمع إلى المفرد"


def rtl_fix(s):
    """Force RTL base direction for cells not starting with a strong RTL letter."""
    for ch in s:
        if ch.isspace():
            continue
        if "؀" <= ch <= "ۿ" or "ݐ" <= ch <= "ݿ":
            return s
        return RLM + s
    return s


def load(name):
    with open(os.path.join(ROOT, name), encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))[1:]  # drop header
    # strip display-only RLM marks so they don't pollute keys/values
    return [[c.replace(RLM, "") for c in row] for row in rows]


def disp(s):
    """Display form: RLM removed, whitespace collapsed (annotations kept)."""
    return re.sub(r"\s+", " ", s.replace(RLM, "")).strip()


def key_norm(s):
    """Exact match key: drop newline/parenthetical annotations, keep diacritics."""
    return re.sub(r"\s+", " ", re.split(r"[\n(]", s)[0]).strip()


def loose(s):
    """Reconciliation key: like key_norm but also ignore the 'noise' diacritics
    (sukun, fatha, tatweel) that the two books spell inconsistently. Damma and
    kasra are KEPT so e.g. فُعلان (damma) never fuses with فِعلان (kasra)."""
    return key_norm(s).translate(str.maketrans("", "", SUKUN + FATHA + TATWEEL))


s2p = load(S2P_NAME + ".csv")  # [sing, type, plur, ex]
p2s = load(P2S_NAME + ".csv")  # [plur, sing, type, ex]

# ---- stage 1: exact merge on (singular, plural, type) -----------------------
rec = {}


def get(key):
    if key not in rec:
        rec[key] = {
            "sing_s2p": "", "plur_s2p": "", "type_s2p": "", "ex_s2p": [],
            "sing_p2s": "", "plur_p2s": "", "type_p2s": "", "ex_p2s": [],
        }
    return rec[key]


for sing, typ, plur, ex in s2p:
    r = get((key_norm(sing), key_norm(plur), key_norm(typ)))
    r["sing_s2p"], r["plur_s2p"], r["type_s2p"] = disp(sing), disp(plur), disp(typ)
    if ex.strip():
        r["ex_s2p"].append(disp(ex))

for plur, sing, typ, ex in p2s:
    r = get((key_norm(sing), key_norm(plur), key_norm(typ)))
    r["sing_p2s"], r["plur_p2s"], r["type_p2s"] = disp(sing), disp(plur), disp(typ)
    if ex.strip():
        r["ex_p2s"].append(disp(ex))


def in_s2p(r):
    return bool(r["sing_s2p"] or r["plur_s2p"] or r["ex_s2p"] or r["type_s2p"])


def in_p2s(r):
    return bool(r["sing_p2s"] or r["plur_p2s"] or r["ex_p2s"] or r["type_p2s"])


# ---- stage 2: reconcile the single-source leftovers -------------------------
# Group leftovers by a diacritic-insensitive (singular, plural) key and fuse an
# s2p-only record with its complementary p2s-only record.
def sole_sing(r):
    return r["sing_s2p"] or r["sing_p2s"]


def sole_plur(r):
    return r["plur_s2p"] or r["plur_p2s"]


groups = {}
for k, r in rec.items():
    if in_s2p(r) and in_p2s(r):
        continue  # already in both → not a leftover
    groups.setdefault((loose(sole_sing(r)), loose(sole_plur(r))), []).append(k)

for keys in groups.values():
    s_keys = [k for k in keys if in_s2p(rec[k]) and not in_p2s(rec[k])]
    p_keys = [k for k in keys if in_p2s(rec[k]) and not in_s2p(rec[k])]
    if not (s_keys and p_keys):
        continue  # no complement to fuse with
    base_k, drop_k = s_keys[0], p_keys[0]
    b, d = rec[base_k], rec[drop_k]
    b["sing_p2s"], b["plur_p2s"], b["type_p2s"], b["ex_p2s"] = (
        d["sing_p2s"], d["plur_p2s"], d["type_p2s"], d["ex_p2s"])
    del rec[drop_k]

# ---- build the differences note + output rows -------------------------------
def longer(a, b):
    return a if len(a) >= len(b) else b


def differences(r):
    notes = []
    s_in, p_in = in_s2p(r), in_p2s(r)
    if s_in and not p_in:
        return f"ورد في «{S2P_NAME}» فقط."
    if p_in and not s_in:
        return f"ورد في «{P2S_NAME}» فقط."
    if r["type_s2p"] != r["type_p2s"]:
        notes.append(
            f"النوع مختلف: «{r['type_s2p']}» في «{S2P_NAME}»، "
            f"و«{r['type_p2s']}» في «{P2S_NAME}».")
    if r["sing_s2p"] != r["sing_p2s"]:
        notes.append(
            f"ضبط المفرد مختلف: «{r['sing_s2p']}» مقابل «{r['sing_p2s']}».")
    if r["plur_s2p"] != r["plur_p2s"]:
        notes.append(
            f"صيغة الجمع مختلفة: «{r['plur_s2p']}» مقابل «{r['plur_p2s']}».")
    return " ".join(notes)


rows = []
for r in rec.values():
    sing = longer(r["sing_s2p"], r["sing_p2s"])
    plur = longer(r["plur_s2p"], r["plur_p2s"])
    if r["type_s2p"] and r["type_p2s"] and r["type_s2p"] != r["type_p2s"]:
        typ = "، ".join([r["type_s2p"], r["type_p2s"]])
    else:
        typ = r["type_s2p"] or r["type_p2s"]
    # examples: if the two tables differ, keep the more complete one and note
    # the other table's version in الفروق.
    ex_a = "\n".join(dict.fromkeys(r["ex_s2p"]))
    ex_b = "\n".join(dict.fromkeys(r["ex_p2s"]))
    ex_note = ""
    if ex_a and ex_b and ex_a != ex_b:
        if len(ex_a) >= len(ex_b):
            examples, alt_name, alt = ex_a, P2S_NAME, ex_b
            keep_name = S2P_NAME
        else:
            examples, alt_name, alt = ex_b, S2P_NAME, ex_a
            keep_name = P2S_NAME
        ex_note = (f"الأمثلة مختلفة: المثبَت من «{keep_name}»؛ "
                   f"وأمثلة «{alt_name}»: {alt}")
    else:
        examples = ex_a or ex_b

    note = " ".join(p for p in (differences(r), ex_note) if p)
    rows.append([sing, typ, plur, examples, note])

rows.sort(key=lambda x: (x[0], x[2], x[1]))

headers = ["المفرد", "النوع", "الجمع", "الأمثلة", "الفروق"]

print(f"merged rows: {len(rows)}")
print(f"  rows with a noted difference: {sum(1 for x in rows if x[4])}")
kinds = Counter()
for x in rows:
    d = x[4]
    if not d:
        kinds["متطابق (لا فروق)"] += 1
    elif "فقط" in d:
        kinds["ورد في جدول واحد فقط"] += 1
    else:
        if "النوع مختلف" in d:
            kinds["اختلاف النوع"] += 1
        if "ضبط المفرد" in d:
            kinds["اختلاف ضبط المفرد"] += 1
        if "صيغة الجمع" in d:
            kinds["اختلاف صيغة الجمع"] += 1
        if "الأمثلة مختلفة" in d:
            kinds["اختلاف الأمثلة"] += 1
for k, v in kinds.most_common():
    print(f"    {k}: {v}")

base = os.path.join(ROOT, "الجدول الموحد")

# CSV
with open(base + ".csv", "w", encoding="utf-8-sig", newline="") as f:
    w = csv.writer(f)
    w.writerow(headers)
    w.writerows([rtl_fix(c) for c in row] for row in rows)

# XLSX
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "الجدول الموحد"
ws.sheet_view.rightToLeft = True
ws.append(headers)
for row in rows:
    ws.append([rtl_fix(c) for c in row])
bold = Font(bold=True)
for cell in ws[1]:
    cell.font = bold
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, readingOrder=2)
wrap = Alignment(vertical="top", wrap_text=True, readingOrder=2)
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = wrap
for i, wdt in enumerate([16, 12, 16, 80, 45], start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = wdt
ws.freeze_panes = "A2"
wb.save(base + ".xlsx")

print(f"-> {base}.csv")
print(f"-> {base}.xlsx")
