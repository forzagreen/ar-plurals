#!/usr/bin/env python3
"""Extract the two big plural-form tables from the Word chapters to xlsx + csv."""
import csv
import os
import zipfile
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.styles import Alignment, Font

W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
NS = {"w": W}

# archival: the .docx chapters and the extracted tables live in sources/.
HERE = os.path.dirname(os.path.abspath(__file__))
SOURCES = os.path.join(os.path.dirname(HERE), "sources")
OUT_DIR = SOURCES

RLM = "‏"  # right-to-left mark


def rtl_fix(s):
    """Force RTL base direction for cells that don't start with a strong RTL
    letter (e.g. cells beginning with '(' or '-'), so parentheses and trailing
    '...' render on the correct side."""
    for ch in s:
        if ch.isspace():
            continue
        # Arabic / Arabic Supplement ranges = strong RTL → already fine
        if "؀" <= ch <= "ۿ" or "ݐ" <= ch <= "ݿ":
            return s
        return RLM + s
    return s


JOBS = [
    {
        "docx": "3. من المفرد إلى الجمع.docx",
        "out": "من المفرد إلى الجمع",
        "headers": ["المفرد", "النوع", "الجمع", "الأمثلة"],
    },
    {
        "docx": "5. من الجمع إلى المفرد.docx",
        "out": "من الجمع إلى المفرد",
        "headers": ["الجمع", "المفرد", "النوع", "الأمثلة"],
    },
]


def cell_text(tc):
    """Cell text: paragraphs joined by newline, runs concatenated."""
    paras = []
    for p in tc.findall("w:p", NS):
        paras.append("".join(t.text or "" for t in p.findall(".//w:t", NS)))
    return "\n".join(paras).strip()


def read_rows(docx_path):
    z = zipfile.ZipFile(docx_path)
    root = ET.fromstring(z.read("word/document.xml"))
    body = root.find("w:body", NS)
    rows = []
    for tbl in body.findall(".//w:tbl", NS):
        for tr in tbl.findall("w:tr", NS):
            cells = [cell_text(tc) for tc in tr.findall("w:tc", NS)]
            # pad/trim to 4 columns
            cells = (cells + [""] * 4)[:4]
            rows.append(cells)
    return rows


def forward_fill(rows):
    """Fill blanks in the three grouping columns (all but the last) from above."""
    last = [None, None, None]
    for row in rows:
        for c in range(3):
            if row[c]:
                last[c] = row[c]
            elif last[c] is not None:
                row[c] = last[c]
    return rows


def write_csv(path, headers, rows):
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows([rtl_fix(c) for c in row] for row in rows)


def write_xlsx(path, headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الجدول"
    ws.sheet_view.rightToLeft = True
    ws.append(headers)
    for row in rows:
        ws.append([rtl_fix(c) for c in row])
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
    wrap = Alignment(vertical="top", wrap_text=True, readingOrder=2)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    widths = [16, 10, 16, 90]
    for i, wdt in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = wdt
    ws.freeze_panes = "A2"
    wb.save(path)


def main():
    for job in JOBS:
        docx_path = os.path.join(SOURCES, job["docx"])
        rows = forward_fill(read_rows(docx_path))
        print(f"{job['docx']}: {len(rows)} data rows")
        base = os.path.join(OUT_DIR, job["out"])
        write_csv(base + ".csv", job["headers"], rows)
        write_xlsx(base + ".xlsx", job["headers"], rows)
        print(f"  -> {base}.csv")
        print(f"  -> {base}.xlsx")


if __name__ == "__main__":
    main()
